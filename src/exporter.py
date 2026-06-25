import io
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_BG = "1F4E79"
HEADER_FG = "FFFFFF"
ALT_ROW_BG = "D6E4F0"
KPI_BG = "E8F4FD"
ACCENT = "2E86AB"

def _thin_border():
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)

def _style_header_row(ws, row, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color=HEADER_FG, name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()

def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

def export_cleaned_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cleaned Data"
    headers = list(df.columns)
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j)
            cell.value = None if pd.isna(val) else val
            cell.font = Font(name="Arial", size=10)
            cell.border = _thin_border()
            if i % 2 == 0:
                cell.fill = PatternFill("solid", start_color=ALT_ROW_BG)
    ws.freeze_panes = "A2"
    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def export_summary_excel(df, kpis, status_df, top_items_df):
    wb = Workbook()
    ws_kpi = wb.active
    ws_kpi.title = "KPI Summary"
    ws_kpi.column_dimensions["A"].width = 28
    ws_kpi.column_dimensions["B"].width = 20
    ws_kpi["A1"] = "Logistics KPI Summary"
    ws_kpi["A1"].font = Font(bold=True, size=14, color=ACCENT, name="Arial")
    ws_kpi["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_kpi["A2"].font = Font(italic=True, size=9, color="888888", name="Arial")
    kpi_labels = {
        "total_orders": "Total Orders",
        "total_quantity": "Total Quantity",
        "total_revenue": "Total Revenue ($)",
        "unique_items": "Unique Items",
        "unique_customers": "Unique Customers",
    }
    for r, (key, label) in enumerate(kpi_labels.items(), start=4):
        a = ws_kpi.cell(row=r, column=1, value=label)
        b = ws_kpi.cell(row=r, column=2, value=kpis.get(key, "N/A"))
        for cell in (a, b):
            cell.font = Font(name="Arial", size=11)
            cell.fill = PatternFill("solid", start_color=KPI_BG)
            cell.border = _thin_border()
        a.font = Font(name="Arial", size=11, bold=True)
    if not status_df.empty:
        ws_s = wb.create_sheet("Status Breakdown")
        ws_s.append(list(status_df.columns))
        _style_header_row(ws_s, 1, len(status_df.columns))
        for row in status_df.itertuples(index=False):
            ws_s.append(list(row))
        _auto_width(ws_s)
    if not top_items_df.empty:
        ws_t = wb.create_sheet("Top Items")
        ws_t.append(list(top_items_df.columns))
        _style_header_row(ws_t, 1, len(top_items_df.columns))
        for row in top_items_df.itertuples(index=False):
            ws_t.append(list(row))
        _auto_width(ws_t)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
